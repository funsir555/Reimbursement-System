#!/usr/bin/env python3
"""
Runtime helper for interactive PowerShell wrappers.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from excel_mysql_common import DEFAULT_APP_YML, ExcelToolError, open_connection


TYPE_HINT_PATTERN = re.compile(
    r"(char|text|int|decimal|numeric|date|datetime|timestamp|time|double|float|bool|tinyint|bigint|not null|null|default)",
    re.IGNORECASE,
)


def ensure_dependencies() -> tuple[Any, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError as exc:
        raise SystemExit("Missing dependency 'openpyxl'. Run: pip install openpyxl pymysql") from exc
    try:
        import pymysql  # type: ignore
    except ModuleNotFoundError as exc:
        raise SystemExit("Missing dependency 'pymysql'. Run: pip install openpyxl pymysql") from exc
    return openpyxl, pymysql


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Support helpers for PowerShell Excel tools.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    detect_parser = subparsers.add_parser("detect-template", help="Detect Excel import row mode.")
    detect_parser.add_argument("--file", required=True, help="Path to the .xlsx file.")
    detect_parser.add_argument("--sheet", help="Worksheet name. Defaults to the first worksheet.")

    key_parser = subparsers.add_parser("resolve-keys", help="Resolve duplicate-check key columns.")
    key_parser.add_argument("--table", required=True, help="Target table name.")
    key_parser.add_argument(
        "--app-yml",
        default=str(DEFAULT_APP_YML),
        help="Path to application.yml. Default: backend auth-service datasource config.",
    )
    return parser.parse_args()


def non_empty_ratio(values: list[Any]) -> float:
    if not values:
        return 0.0
    return sum(1 for value in values if value not in (None, "")) / len(values)


def detect_template_mode(openpyxl: Any, file_path: Path, sheet_name: str | None) -> dict[str, Any]:
    if not file_path.exists():
        raise ExcelToolError(f"Excel file not found: {file_path}")
    if file_path.suffix.lower() != ".xlsx":
        raise ExcelToolError("Only .xlsx files are supported.")

    workbook = openpyxl.load_workbook(filename=file_path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ExcelToolError(f"Worksheet not found: {sheet_name}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    row1 = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    headers = [str(value).strip() for value in row1 if value is not None and str(value).strip() != ""]
    if not headers:
        raise ExcelToolError("Header row is empty.")

    header_count = len(headers)
    row2 = [sheet.cell(row=2, column=index).value for index in range(1, header_count + 1)]
    row3 = [sheet.cell(row=3, column=index).value for index in range(1, header_count + 1)]

    row2_ratio = non_empty_ratio(row2)
    row3_ratio = non_empty_ratio(row3)
    row3_has_type_hint = any(TYPE_HINT_PATTERN.search(str(value)) for value in row3 if value not in (None, ""))

    is_template = row2_ratio >= 0.6 and row3_ratio >= 0.6 and row3_has_type_hint
    return {
        "header_row": 1,
        "start_row": 4 if is_template else 2,
        "mode": "template" if is_template else "simple",
        "reason": (
            "row 2/3 contain comment and constraint metadata"
            if is_template
            else "row 2/3 do not look like exported template metadata"
        ),
    }


def resolve_key_columns(cursor: Any, table_name: str) -> dict[str, Any]:
    cursor.execute(
        """
        SELECT COLUMN_NAME, EXTRA
        FROM information_schema.COLUMNS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (table_name,),
    )
    column_rows = cursor.fetchall()
    if not column_rows:
        raise ExcelToolError(f"Target table not found: {table_name}")
    column_extra = {row["COLUMN_NAME"]: (row["EXTRA"] or "").lower() for row in column_rows}

    cursor.execute(
        """
        SELECT
            INDEX_NAME,
            NON_UNIQUE,
            SEQ_IN_INDEX,
            COLUMN_NAME
        FROM information_schema.STATISTICS
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        ORDER BY
            CASE WHEN INDEX_NAME = 'PRIMARY' THEN 0 ELSE 1 END,
            INDEX_NAME,
            SEQ_IN_INDEX
        """,
        (table_name,),
    )
    rows = cursor.fetchall()
    if not rows:
        raise ExcelToolError(f"Target table not found or has no indexes: {table_name}")

    indexes: dict[str, list[str]] = {}
    unique_indexes: list[dict[str, Any]] = []
    for row in rows:
        index_name = row["INDEX_NAME"]
        indexes.setdefault(index_name, []).append(row["COLUMN_NAME"])
    for index_name, columns in indexes.items():
        non_unique = next(item["NON_UNIQUE"] for item in rows if item["INDEX_NAME"] == index_name)
        if non_unique == 0:
            unique_indexes.append(
                {
                    "index_name": index_name,
                    "columns": columns,
                    "has_auto_increment": any("auto_increment" in column_extra.get(column, "") for column in columns),
                }
            )

    primary = next((item for item in unique_indexes if item["index_name"] == "PRIMARY"), None)
    if primary and not primary["has_auto_increment"]:
        return {
            "status": "selected",
            "source": "primary_key",
            "key_columns": primary["columns"],
            "candidates": [primary],
        }

    non_primary_unique_indexes = [item for item in unique_indexes if item["index_name"] != "PRIMARY" and not item["has_auto_increment"]]
    if not primary and not non_primary_unique_indexes:
        return {"status": "missing", "source": "none", "key_columns": [], "candidates": []}

    if primary and not non_primary_unique_indexes:
        return {
            "status": "missing",
            "source": "auto_increment_primary_key_only",
            "key_columns": [],
            "candidates": [primary],
        }

    if len(non_primary_unique_indexes) == 1:
        return {
            "status": "selected",
            "source": "unique_index",
            "key_columns": non_primary_unique_indexes[0]["columns"],
            "candidates": non_primary_unique_indexes,
        }

    single_column_indexes = [item for item in non_primary_unique_indexes if len(item["columns"]) == 1]
    if len(single_column_indexes) == 1:
        return {
            "status": "selected",
            "source": "single_unique_index",
            "key_columns": single_column_indexes[0]["columns"],
            "candidates": non_primary_unique_indexes,
        }

    return {
        "status": "ambiguous",
        "source": "multiple_unique_indexes",
        "key_columns": [],
        "candidates": non_primary_unique_indexes,
    }


def command_detect_template(args: argparse.Namespace) -> int:
    openpyxl, _ = ensure_dependencies()
    result = detect_template_mode(openpyxl, Path(args.file).expanduser().resolve(), args.sheet)
    print(json.dumps(result, ensure_ascii=False))
    return 0


def command_resolve_keys(args: argparse.Namespace) -> int:
    _, pymysql = ensure_dependencies()
    connection = None
    try:
        connection = open_connection(pymysql, Path(args.app_yml).expanduser().resolve())
        with connection.cursor() as cursor:
            result = resolve_key_columns(cursor, args.table)
        print(json.dumps(result, ensure_ascii=False))
        return 0
    finally:
        if connection is not None:
            connection.close()


def main() -> int:
    args = parse_args()
    try:
        if args.command == "detect-template":
            return command_detect_template(args)
        if args.command == "resolve-keys":
            return command_resolve_keys(args)
        raise ExcelToolError(f"Unsupported command: {args.command}")
    except ExcelToolError as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1
    except Exception as exc:  # pragma: no cover - defensive CLI fallback
        print(f"[ERROR] Unexpected failure: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
