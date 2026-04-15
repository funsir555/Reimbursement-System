#!/usr/bin/env python3
"""
Generic Excel importer for finex_db.

Recommended setup on Windows:
    python -m venv .venv
    .\\.venv\\Scripts\\python.exe -m pip install --upgrade pip
    .\\.venv\\Scripts\\python.exe -m pip install openpyxl pymysql

Examples:
    .\\.venv\\Scripts\\python.exe import_excel_to_mysql.py ^
        --file C:\\data\\customer.xlsx ^
        --table gl_Customer ^
        --key-columns cCusCode ^
        --dry-run

    .\\.venv\\Scripts\\python.exe import_excel_to_mysql.py ^
        --file C:\\data\\customer_template.xlsx ^
        --table gl_Customer ^
        --key-columns cCusCode ^
        --header-row 1 ^
        --start-row 4 ^
        --dry-run

Rules:
    - The first header row must use database column names.
    - Templates exported by export_table_template.py keep row 1 as headers,
      row 2 as Chinese comments, row 3 as column constraints, and row 4+ as data.
    - Existing keys cause the import to fail immediately.
    - DATE / DATETIME / TIMESTAMP columns are validated and normalized.
    - DECIMAL(18,2) columns are treated as amount columns by default.
    - Suspected garbled text values are rejected.
    - Use --dry-run before a real import.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Sequence

from excel_mysql_common import (
    DEFAULT_APP_YML,
    ColumnMeta,
    ExcelToolError,
    fetch_columns,
    fetch_regex_check_constraints,
    is_system_managed_column,
    open_connection,
    quote_identifier,
    split_csv_arg,
)

DATE_INPUT_FORMATS = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M:%S",
)
TEXT_DATA_TYPES = {"char", "varchar", "text", "mediumtext", "longtext"}
DATE_DATA_TYPES = {"date", "datetime", "timestamp"}
AMOUNT_QUANT = Decimal("0.01")
SUSPICIOUS_TEXT_PATTERNS = (
    "????",
    "???",
    "??",
    "\ufffd",
    "锟",
    "�",
)
INSERT_BATCH_SIZE = 200
TEXT_LENGTH_PATTERN = re.compile(r"^(?:var)?char\((\d+)\)$")


@dataclass
class ValidationError(ExcelToolError):
    row_number: int
    column_name: str
    raw_value: Any
    reason: str

    def __str__(self) -> str:
        return (
            f"第 {self.row_number} 行，列“{self.column_name}”，原值={self.raw_value!r}：{self.reason}"
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import .xlsx rows into a MySQL table. For templates exported by "
        "export_table_template.py, use --header-row 1 --start-row 4."
    )
    parser.add_argument("--file", required=True, help="Path to the .xlsx file.")
    parser.add_argument("--table", required=True, help="Target table name.")
    parser.add_argument("--sheet", help="Worksheet name. Defaults to the first worksheet.")
    parser.add_argument(
        "--key-columns",
        required=True,
        help="Comma-separated unique key columns used for duplicate checks.",
    )
    parser.add_argument("--header-row", type=int, default=1, help="Header row number. Default: 1.")
    parser.add_argument("--start-row", type=int, default=2, help="Data start row number. Default: 2.")
    parser.add_argument("--dry-run", action="store_true", help="Validate only, do not write to DB.")
    parser.add_argument(
        "--date-columns",
        default="",
        help="Extra comma-separated columns forced to use date validation.",
    )
    parser.add_argument(
        "--amount-columns",
        default="",
        help="Extra comma-separated columns forced to use amount validation.",
    )
    parser.add_argument(
        "--text-columns",
        default="",
        help="Extra comma-separated columns forced to use strict text/garbled validation.",
    )
    parser.add_argument(
        "--app-yml",
        default=str(DEFAULT_APP_YML),
        help="Path to application.yml. Default: backend auth-service datasource config.",
    )
    args = parser.parse_args()
    if args.header_row < 1 or args.start_row < 1:
        parser.error("--header-row and --start-row must be positive integers.")
    if args.start_row <= args.header_row:
        parser.error("--start-row must be greater than --header-row.")
    return args


def ensure_dependencies() -> tuple[Any, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError as exc:
        raise SystemExit("缺少依赖 openpyxl，请先执行：pip install openpyxl pymysql") from exc
    try:
        import pymysql  # type: ignore
    except ModuleNotFoundError as exc:
        raise SystemExit("缺少依赖 pymysql，请先执行：pip install openpyxl pymysql") from exc
    return openpyxl, pymysql


def load_workbook(openpyxl: Any, file_path: Path, sheet_name: str | None) -> Any:
    if not file_path.exists():
        raise ExcelToolError(f"Excel 文件不存在：{file_path}")
    workbook = openpyxl.load_workbook(filename=file_path, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ExcelToolError(f"工作表不存在：{sheet_name}")
        return workbook[sheet_name]
    return workbook[workbook.sheetnames[0]]


def read_headers(sheet: Any, header_row: int) -> list[str]:
    headers = []
    seen = set()
    header_cells = next(sheet.iter_rows(min_row=header_row, max_row=header_row))
    for cell in header_cells:
        value = cell.value
        if value is None or str(value).strip() == "":
            continue
        header = str(value).strip()
        if header in seen:
            raise ExcelToolError(f"Excel 表头重复：{header}")
        seen.add(header)
        headers.append(header)
    if not headers:
        raise ExcelToolError("Excel 表头行为空。")
    return headers


def validate_headers(headers: Sequence[str], columns: dict[str, ColumnMeta], key_columns: Sequence[str]) -> None:
    db_columns = set(columns)
    extra_headers = [header for header in headers if header not in db_columns]
    if extra_headers:
        raise ExcelToolError(f"Excel 包含数据库中不存在的列：{', '.join(extra_headers)}")
    missing_required = []
    for name, meta in columns.items():
        if name in headers:
            continue
        if meta.is_nullable or meta.column_default is not None or "auto_increment" in meta.extra or is_system_managed_column(meta):
            continue
        missing_required.append(name)
    if missing_required:
        raise ExcelToolError(
            "Excel 缺少数据库必填列：" + ", ".join(missing_required)
        )
    missing_keys = [column for column in key_columns if column not in headers]
    if missing_keys:
        raise ExcelToolError("Excel 表头缺少唯一键列：" + ", ".join(missing_keys))


def is_blank_row(row_values: Sequence[Any]) -> bool:
    return all(value is None or str(value).strip() == "" for value in row_values)


def parse_date_value(value: Any, meta: ColumnMeta, row_number: int, column_name: str) -> Any:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, dt.datetime):
        if meta.data_type == "date":
            return value.date()
        return value.replace(microsecond=0)
    if isinstance(value, dt.date):
        if meta.data_type == "date":
            return value
        return dt.datetime.combine(value, dt.time.min)
    if isinstance(value, str):
        text = value.strip()
        for fmt in DATE_INPUT_FORMATS:
            try:
                parsed = dt.datetime.strptime(text, fmt)
                if meta.data_type == "date":
                    return parsed.date()
                return parsed
            except ValueError:
                continue
    raise ValidationError(row_number, column_name, value, "日期格式不正确")


def parse_amount_value(value: Any, row_number: int, column_name: str) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, bool):
        raise ValidationError(row_number, column_name, value, "布尔值不是合法金额")
    if isinstance(value, str):
        text = value.strip()
        if "," in text:
            raise ValidationError(row_number, column_name, value, "金额不支持千分位逗号格式")
        candidate = text
    else:
        candidate = str(value)
    try:
        decimal_value = Decimal(candidate)
    except InvalidOperation as exc:
        raise ValidationError(row_number, column_name, value, "金额不是合法数字") from exc
    if decimal_value.as_tuple().exponent < -2:
        raise ValidationError(row_number, column_name, value, "金额小数位不能超过 2 位")
    return decimal_value.quantize(AMOUNT_QUANT)

def parse_text_length_limit(column_type: str) -> int | None:
    match = TEXT_LENGTH_PATTERN.fullmatch(column_type.strip().lower())
    if not match:
        return None
    return int(match.group(1))


def check_text_value(value: Any, row_number: int, column_name: str, strict: bool, max_length: int | None) -> str | None:
    if value is None:
        return None
    text = str(value)
    if text.strip() == "":
        return None
    for marker in SUSPICIOUS_TEXT_PATTERNS:
        if marker in text:
            raise ValidationError(row_number, column_name, value, "疑似乱码文本")
    if strict and text.strip() in {"?", "??", "???", "????", "NULL", "null"}:
        raise ValidationError(row_number, column_name, value, "不允许使用占位文本")
    if max_length is not None and len(text.strip()) > max_length:
        raise ValidationError(row_number, column_name, value, f"文本长度超过上限 {max_length}")
    return text


def normalize_row(
    row_number: int,
    row_map: dict[str, Any],
    columns: dict[str, ColumnMeta],
    date_columns: set[str],
    amount_columns: set[str],
    strict_text_columns: set[str],
) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for column_name, value in row_map.items():
        meta = columns[column_name]
        if value is None or (isinstance(value, str) and value.strip() == ""):
            normalized[column_name] = None
            continue
        if column_name in date_columns:
            normalized[column_name] = parse_date_value(value, meta, row_number, column_name)
            continue
        if column_name in amount_columns:
            normalized[column_name] = parse_amount_value(value, row_number, column_name)
            continue
        if meta.data_type in TEXT_DATA_TYPES or column_name in strict_text_columns:
            normalized[column_name] = check_text_value(
                value,
                row_number,
                column_name,
                column_name in strict_text_columns,
                parse_text_length_limit(meta.column_type),
            )
            continue
        normalized[column_name] = value
    return normalized


def resolve_insert_columns(
    headers: Sequence[str],
    row_map: dict[str, Any],
    columns: dict[str, ColumnMeta],
) -> list[str]:
    insert_columns: list[str] = []
    for column_name in headers:
        value = row_map.get(column_name)
        if value is None and is_system_managed_column(columns[column_name]):
            continue
        insert_columns.append(column_name)
    return insert_columns


def build_key_tuple(row_map: dict[str, Any], key_columns: Sequence[str], row_number: int) -> tuple[Any, ...]:
    values = []
    for column in key_columns:
        value = row_map.get(column)
        if value is None or str(value).strip() == "":
            raise ValidationError(row_number, column, value, "key column cannot be empty")
        values.append(value)
    return tuple(values)


def detect_duplicate_keys_in_excel(key_rows: Iterable[tuple[int, tuple[Any, ...]]]) -> None:
    seen: dict[tuple[Any, ...], int] = {}
    for row_number, key_tuple in key_rows:
        if key_tuple in seen:
            raise ExcelToolError(
                f"Excel 中发现重复唯一键：首次出现在第 {seen[key_tuple]} 行，重复出现在第 {row_number} 行，键值={key_tuple!r}"
            )
        seen[key_tuple] = row_number


def detect_existing_keys(cursor: Any, table_name: str, key_columns: Sequence[str], key_values: Sequence[tuple[Any, ...]]) -> None:
    if not key_values:
        return
    key_expr = ", ".join(quote_identifier(column) for column in key_columns)
    chunk_size = 500
    for start in range(0, len(key_values), chunk_size):
        chunk = key_values[start : start + chunk_size]
        placeholders = ", ".join(["(" + ", ".join(["%s"] * len(key_columns)) + ")"] * len(chunk))
        sql = (
            f"SELECT {key_expr} FROM {quote_identifier(table_name)} "
            f"WHERE ({key_expr}) IN ({placeholders}) LIMIT 1"
        )
        params: list[Any] = []
        for key_tuple in chunk:
            params.extend(key_tuple)
        cursor.execute(sql, params)
        row = cursor.fetchone()
        if row:
            existing_key = tuple(row[column] for column in key_columns)
            raise ExcelToolError(f"数据库中已存在相同唯一键：{existing_key!r}")


def validate_check_constraints(
    cursor: Any,
    table_name: str,
    rows: Sequence[tuple[int, dict[str, Any]]],
) -> None:
    regex_constraints = fetch_regex_check_constraints(cursor, table_name)
    if not regex_constraints:
        return
    compiled_constraints: list[tuple[str, str, re.Pattern[str]]] = []
    for item in regex_constraints:
        try:
            compiled_constraints.append((item.name, item.column_name, re.compile(item.pattern)))
        except re.error:
            continue
    for row_number, row in rows:
        for constraint_name, column_name, pattern in compiled_constraints:
            if column_name not in row:
                continue
            value = row[column_name]
            if value is None:
                continue
            if pattern.fullmatch(str(value)):
                continue
            raise ValidationError(
                row_number,
                column_name,
                value,
                f"违反数据库检查约束 {constraint_name}，要求匹配 {pattern.pattern}",
            )


def insert_rows(
    cursor: Any,
    table_name: str,
    headers: Sequence[str],
    rows: Sequence[dict[str, Any]],
    columns: dict[str, ColumnMeta],
) -> int:
    if not rows:
        return 0
    inserted = 0
    batches: dict[tuple[str, ...], list[dict[str, Any]]] = {}
    for row in rows:
        insert_columns = tuple(resolve_insert_columns(headers, row, columns))
        if not insert_columns:
            raise ExcelToolError("存在无法写入的空行，请检查 Excel 数据。")
        batches.setdefault(insert_columns, []).append(row)

    for insert_columns, grouped_rows in batches.items():
        sql = (
            f"INSERT INTO {quote_identifier(table_name)} "
            f"({', '.join(quote_identifier(column) for column in insert_columns)}) "
            f"VALUES ({', '.join(['%s'] * len(insert_columns))})"
        )
        for start in range(0, len(grouped_rows), INSERT_BATCH_SIZE):
            batch = grouped_rows[start : start + INSERT_BATCH_SIZE]
            params = [tuple(row.get(column) for column in insert_columns) for row in batch]
            cursor.executemany(sql, params)
            inserted += len(batch)
            print(f"[INFO] 已插入 {inserted}/{len(rows)} 行")
    return inserted


def main() -> int:
    args = parse_args()
    openpyxl, pymysql = ensure_dependencies()

    excel_path = Path(args.file).expanduser().resolve()
    app_yml_path = Path(args.app_yml).expanduser().resolve()
    key_columns = split_csv_arg(args.key_columns)
    forced_date_columns = set(split_csv_arg(args.date_columns))
    forced_amount_columns = set(split_csv_arg(args.amount_columns))
    forced_text_columns = set(split_csv_arg(args.text_columns))

    connection = None
    try:
        connection = open_connection(pymysql, app_yml_path)
        with connection.cursor() as cursor:
            columns = fetch_columns(cursor, args.table)
            sheet = load_workbook(openpyxl, excel_path, args.sheet)
            headers = read_headers(sheet, args.header_row)
            validate_headers(headers, columns, key_columns)
            system_managed_header_columns = [
                name for name in headers if is_system_managed_column(columns[name])
            ]

            unknown_forced = (forced_date_columns | forced_amount_columns | forced_text_columns) - set(headers)
            if unknown_forced:
                raise ExcelToolError(
                    "强制校验列未出现在 Excel 表头中："
                    + ", ".join(sorted(unknown_forced))
                )

            date_columns = {
                name for name in headers if columns[name].data_type in DATE_DATA_TYPES
            } | forced_date_columns
            amount_columns = {
                name for name in headers if columns[name].column_type == "decimal(18,2)"
            } | forced_amount_columns

            normalized_rows: list[dict[str, Any]] = []
            numbered_rows: list[tuple[int, dict[str, Any]]] = []
            excel_keys: list[tuple[int, tuple[Any, ...]]] = []
            total_rows = 0
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=args.start_row, values_only=True),
                start=args.start_row,
            ):
                row_values = list(row[: len(headers)])
                if is_blank_row(row_values):
                    continue
                total_rows += 1
                row_map = {headers[idx]: row_values[idx] if idx < len(row_values) else None for idx in range(len(headers))}
                normalized = normalize_row(
                    row_number=row_number,
                    row_map=row_map,
                    columns=columns,
                    date_columns=date_columns,
                    amount_columns=amount_columns,
                    strict_text_columns=forced_text_columns,
                )
                key_tuple = build_key_tuple(normalized, key_columns, row_number)
                excel_keys.append((row_number, key_tuple))
                normalized_rows.append(normalized)
                numbered_rows.append((row_number, normalized))

            if total_rows == 0:
                raise ExcelToolError("Excel 中未找到可导入的数据行。")

            detect_duplicate_keys_in_excel(excel_keys)
            detect_existing_keys(cursor, args.table, key_columns, [item[1] for item in excel_keys])
            validate_check_constraints(cursor, args.table, numbered_rows)

            print(f"[INFO] Excel 校验通过行数：{len(normalized_rows)}")
            print(f"[INFO] 日期列：{sorted(date_columns)}")
            print(f"[INFO] 金额列：{sorted(amount_columns)}")
            print(f"[INFO] 严格文本列：{sorted(forced_text_columns)}")
            print(f"[INFO] 已自动忽略系统托管列（空值时不写入）：{sorted(system_managed_header_columns)}")

            if args.dry_run:
                print("[INFO] Dry-run 校验成功，未写入数据库。")
                connection.rollback()
                return 0

            inserted = insert_rows(cursor, args.table, headers, normalized_rows, columns)
            connection.commit()
            print(f"[INFO] 导入完成，成功写入 {inserted} 行。")
            return 0
    except ValidationError as exc:
        if connection is not None:
            connection.rollback()
        print(f"[ERROR] 校验失败：{exc}", file=sys.stderr)
        return 1
    except ExcelToolError as exc:
        if connection is not None:
            connection.rollback()
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1
    except Exception as exc:  # pragma: no cover - defensive CLI fallback
        if connection is not None:
            connection.rollback()
        print(f"[ERROR] Unexpected failure: {exc}", file=sys.stderr)
        return 1
    finally:
        if connection is not None:
            connection.close()


if __name__ == "__main__":
    raise SystemExit(main())
