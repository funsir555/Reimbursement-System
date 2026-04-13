#!/usr/bin/env python3
"""
Export a single-table Excel import template for finex_db.

Recommended setup on Windows:
    python -m venv .venv
    .\\.venv\\Scripts\\python.exe -m pip install --upgrade pip
    .\\.venv\\Scripts\\python.exe -m pip install openpyxl pymysql

Examples:
    .\\.venv\\Scripts\\python.exe export_table_template.py ^
        --table gl_Customer

    .\\.venv\\Scripts\\python.exe export_table_template.py ^
        --table pm_document_expense_detail ^
        --output C:\\data\\pm_document_expense_detail.xlsx

The exported workbook format is:
    - row 1: database column names
    - row 2: Chinese column comments
    - row 3: type / nullability / default / key info
    - row 4+: blank rows for manual data entry
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

from excel_mysql_common import (
    DEFAULT_APP_YML,
    ExcelToolError,
    fetch_columns,
    format_column_constraint,
    open_connection,
)


DEFAULT_TEMPLATE_DIR = Path(__file__).resolve().parent / "templates"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export a single-table .xlsx template using database column metadata."
    )
    parser.add_argument("--table", required=True, help="Target table name.")
    parser.add_argument(
        "--output",
        help="Output .xlsx path. Default: backend/sql/templates/<table>.xlsx",
    )
    parser.add_argument(
        "--app-yml",
        default=str(DEFAULT_APP_YML),
        help="Path to application.yml. Default: backend auth-service datasource config.",
    )
    include_group = parser.add_mutually_exclusive_group()
    include_group.add_argument(
        "--include-meta",
        dest="include_meta",
        action="store_true",
        default=True,
        help="Include row 2 comments and row 3 column constraints. Default: enabled.",
    )
    include_group.add_argument(
        "--no-include-meta",
        dest="include_meta",
        action="store_false",
        help="Export only row 1 field names.",
    )
    return parser.parse_args()


def ensure_dependencies() -> tuple[Any, Any]:
    try:
        import openpyxl  # type: ignore
    except ModuleNotFoundError as exc:
        raise SystemExit("Missing dependency 'openpyxl'. Run: pip install openpyxl pymysql") from exc
    try:
        import pymysql  # type: ignore
    except ModuleNotFoundError as exc:
        raise SystemExit("Missing dependency 'pymysql'. Run: pip install openpyxl pymysql") from exc
    return openpyxl, pymysql


def resolve_output_path(table_name: str, output_arg: str | None) -> Path:
    if output_arg:
        output_path = Path(output_arg).expanduser().resolve()
    else:
        output_path = (DEFAULT_TEMPLATE_DIR / f"{table_name}.xlsx").resolve()
    if output_path.suffix.lower() != ".xlsx":
        raise ExcelToolError("--output must end with .xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    return output_path


def autosize_columns(sheet: Any, headers: list[str], comments: list[str], constraints: list[str], include_meta: bool) -> None:
    rows = [headers]
    if include_meta:
        rows.append(comments)
        rows.append(constraints)
    for idx, values in enumerate(zip(*rows), start=1):
        max_len = max(len(str(value or "")) for value in values)
        sheet.column_dimensions[openpyxl_column_letter(idx)].width = min(max(max_len + 4, 14), 48)


def openpyxl_column_letter(index: int) -> str:
    letters = []
    current = index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def write_template(openpyxl: Any, table_name: str, columns: dict[str, Any], output_path: Path, include_meta: bool) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = table_name[:31]

    headers = [meta.name for meta in columns.values()]
    comments = [meta.comment or "" for meta in columns.values()]
    constraints = [format_column_constraint(meta) for meta in columns.values()]

    for column_index, value in enumerate(headers, start=1):
        sheet.cell(row=1, column=column_index, value=value)

    if include_meta:
        for column_index, value in enumerate(comments, start=1):
            sheet.cell(row=2, column=column_index, value=value)
        for column_index, value in enumerate(constraints, start=1):
            sheet.cell(row=3, column=column_index, value=value)
        sheet.freeze_panes = "A4"
    else:
        sheet.freeze_panes = "A2"

    header_fill = openpyxl.styles.PatternFill("solid", fgColor="D9EAF7")
    meta_fill = openpyxl.styles.PatternFill("solid", fgColor="F3F6D8")
    constraint_fill = openpyxl.styles.PatternFill("solid", fgColor="F6E3CF")
    bold_font = openpyxl.styles.Font(bold=True)
    top_align = openpyxl.styles.Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = top_align
    if include_meta:
        for cell in sheet[2]:
            cell.fill = meta_fill
            cell.alignment = top_align
        for cell in sheet[3]:
            cell.fill = constraint_fill
            cell.alignment = top_align

    autosize_columns(sheet, headers, comments, constraints, include_meta)
    workbook.save(output_path)


def main() -> int:
    args = parse_args()
    openpyxl, pymysql = ensure_dependencies()
    app_yml_path = Path(args.app_yml).expanduser().resolve()
    output_path = resolve_output_path(args.table, args.output)

    connection = None
    try:
        connection = open_connection(pymysql, app_yml_path)
        with connection.cursor() as cursor:
            columns = fetch_columns(cursor, args.table)
        write_template(openpyxl, args.table, columns, output_path, args.include_meta)
        print(f"[INFO] Exported template: {output_path}")
        print("[INFO] Row 1=field names, row 2=comments, row 3=constraints, row 4+=data")
        return 0
    except ExcelToolError as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1
    except Exception as exc:  # pragma: no cover - defensive CLI fallback
        print(f"[ERROR] Unexpected failure: {exc}", file=sys.stderr)
        return 1
    finally:
        if connection is not None:
            connection.close()


if __name__ == "__main__":
    raise SystemExit(main())
